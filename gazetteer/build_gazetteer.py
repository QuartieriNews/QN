#!/usr/bin/env python3
"""
build_gazetteer.py — regenerate every derived gazetteer artefact from the master.

    python build_gazetteer.py

Master (edit this):      Rome_Neighbourhood_Gazetteer_EN.xlsx
Translation memory:      it_strings.json          (hand-maintained, English -> Italian)
Generated (never edit):  gazetteer.json
                         prompt_list_publication_zones.txt
                         Mappa_Quartieri_Normalizzata.xlsx

The resolution policy, the guards, the editorial filter and the reject-reason enum are NOT
in the workbook — they are code-level configuration and live in the CONFIG block below.
Change them here, then re-run.

Not touched by this script: venues.json. Venue records are learned at runtime and have a
different lifecycle; regenerating them from a workbook would erase what the system has
learned. See GAZETTEER_README.md section 9.

v1.4 — corrections after external review of the v1.3 package. See CHANGES_v1.4.md:
  * guards split into two tiers; the suspicion guards now run AFTER the venue registry
  * name_coord_mismatch rewritten — the v1.3 rule rejected "Teatro Argentina, Sala
    Squarzina", because location.name is a venue name and not a hierarchical geo field
  * the editorial filter runs explicit_promotion FIRST; in v1.3 it ran last and could
    never be reached for any promotion carrying a proper noun
  * municipality_code added to the stored contract — geo_level = municipality was not
    routable without it
  * verification_status -> geo_verification_status in the double-pass reconciliation
  * normalisation aligned between build and runtime (accents, collapsed spaces)
  * reject reasons carry blocks_publication, so flag-only codes stop being read as rejects
"""

import re
import unicodedata
import json
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font

def norm(s):
    """The single normalisation used for every lookup key, in the build and at runtime.

    v1.3 compared with casefold() here and stripped accents in zone_distribution.py, so the
    build's guarantee that a runtime collision fails the build was not actually true."""
    s = unicodedata.normalize("NFKD", str(s or ""))
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = re.sub(r"[^a-z0-9 ]+", " ", s.casefold())
    return re.sub(r"\s+", " ", s).strip()


VERSION = "1.4"
MASTER = Path("Rome_Neighbourhood_Gazetteer_EN.xlsx")
STRINGS = Path("it_strings.json")
JSON_OUT = Path("gazetteer.json")
PROMPT_OUT = Path("prompt_list_publication_zones.txt")
EXPORT_IT = Path("Mappa_Quartieri_Normalizzata.xlsx")

# ---------------------------------------------------------------- CONFIG ----

RESOLUTION = {
    "target_field": "publication_zone_id",
    "note": (
        "The model resolves straight to the publication zone. The neighbourhood level is "
        "kept for alias matching and is stored whenever it is known deterministically, but "
        "the model is never asked to choose among 440 toponyms."
    ),
    "allowed_values_source": (
        "publication_zones[] where is_page == true, field name_display. "
        "Do not maintain a second copy of this list anywhere."
    ),
    "store": {
        "publication_zone_id": "always, when resolved",
        "neighbourhood_id": (
            "only when known deterministically (venue registry or alias match). "
            "null when the zone came from the model."
        ),
        "publication_zone": "display name, derived — never a key",
        "geo_level": "how precisely the item is placed. See geo_levels below.",
        "municipality_code": (
            "always, when known. Roman numerals I-XV. Derived from the zone's "
            "primary_municipality at zone level; the ONLY routing key at municipality "
            "level, where publication_zone_id is empty by definition."
        ),
    },
    "municipality_code_note": (
        "Added in v1.4. v1.3 declared that a municipality-level item carries an empty zone "
        "and that the layout would use it for the fallback, while storing no field that "
        "said WHICH municipality. That is a rule with no output field — the defect the "
        "project's own principles forbid — and it would have been discovered after the "
        "institutional block had already collected a corpus, i.e. as a migration."
    ),
    "geo_levels": {
        "note": (
            "The events block resolves to neighbourhood or zone. Other blocks do not and "
            "must not be forced to: a municipal commission sits at municipality level by "
            "nature, and inventing a neighbourhood for it would write an invented fact into "
            "the corpus. The field is set by the ingesting block, never by the model."
        ),
        "values": {
            "neighbourhood": "A specific address or venue is known. neighbourhood_id is set.",
            "zone": "The publication zone is known, the finer level is not.",
            "municipality": "Municipality-wide by nature or by fallback. Zone empty.",
            "city": "Rome-wide. Shown on every page or on none — a layout decision.",
        },
        "fallback_ladder": ["neighbourhood", "zone", "municipality", "city"],
        "fallback_note": (
            "A page short of zone-level content pulls municipality-level content to fill "
            "the block. That is a rendering decision taken at layout time from stored "
            "fields — never a re-resolution of the item to a coarser zone, which would "
            "write a false fact into the corpus."
        ),
    },
    "precedence": [
        "human_verified",
        "venue_registry (facebook location.id)",
        "alias_match",
        "ai",
    ],
    "order_of_operations": [
        "0  human_verified on the item -> use it, stop",
        "1  guards, tier A (coordinate validity) -> reject, stop",
        "2  venue registry on location.id -> hit: zone known, skip to 6",
        "3  guards, tier B (suspicion signals) -> reject, stop",
        "4  alias lookup, then the publishable check",
        "5  model, pass 1 and conditionally pass 2",
        "6  validation",
    ],
    "why_guards_are_split": (
        "v1.3 ran all five guards before the venue registry. A venue a person had already "
        "resolved was therefore blocked forever if its records kept arriving with "
        "countryCode = null — which contradicts the registry's whole premise, resolve once "
        "and for all, and made v1.3's own claim that such a record would be 'caught by the "
        "venue registry once reviewed' unreachable. Tier A is about whether the coordinate "
        "is usable at all and must run first. Tier B is suspicion, and a registry hit "
        "makes suspicion moot."
    ),
    "precedence_note": (
        "A human-set value wins over every recomputation and survives a geo_logic_revision "
        "bump. If a later alias or venue lookup disagrees with a human_verified value, "
        "record geo_conflict = true and keep the human value; do not overwrite it silently."
    ),
    "online_events": {
        "note": (
            "v1.3 had this in two incompatible places: the README made it a deterministic "
            "step before the guards, the prompt asked the model to decide, and "
            "SOURCE_DATA_FINDINGS had already measured isOnline as unreliable — true on 1 "
            "of 50, and false on an event literally titled 'Join our online Meditation "
            "Community'. Nobody was actually deciding."
        ),
        "rule": (
            "Deterministic, before the guards, and conservative: an item is online ONLY if "
            "isOnline is true AND there are no coordinates AND there is no location.id. "
            "Anything with a physical venue is treated as physical whatever the flag says, "
            "because a wrong 'online' verdict removes the item from every page silently."
        ),
        "action": "publication_zone_id = empty, geo_level = city, no reject — an online "
                  "event is a valid item with no zone, not a failure",
        "not_the_model": (
            "The model is not asked. It cannot see isOnline, the coordinates or "
            "location.id, so any answer it gives is a guess about a field it was not shown."
        ),
    },
    "empty_is_valid": True,
    "confidence_field": "geo_basis",
    "geo_basis_values": {
        "explicit_address": "The source states a street or venue address",
        "explicit_zone": "The source names the area directly",
        "inferred": "Deduced from context, not stated",
        "unknown": "Not determinable — publication_zone_id must be empty",
    },
    "verification_status_values": {
        "human_verified": "A person decided. Wins over any recomputation.",
        "deterministic_venue": "Resolved from the venue registry on location.id",
        "deterministic_alias": "Exact alias or name match in the gazetteer",
        "ai_agreement": "Two independent passes returned the same zone",
        "ai_single_pass": "One pass, on an explicit address or an explicitly named zone",
        "unresolved": "No zone. The item is in the reject queue.",
    },
    "double_pass": {
        "mode": "conditional",
        "note": (
            "A second independent pass is spent only where the first one is likely to be "
            "wrong. geo_basis is the discriminator and it is already returned by the model. "
            "Applying it everywhere doubles the AI cost on the whole volume to protect a "
            "segment that can be isolated."
        ),
        "single_pass_when": ["explicit_address", "explicit_zone"],
        "second_pass_when": ["inferred", "unknown"],
        "second_pass_also_when": (
            "no venue-registry hit and no street address in the source record — "
            "location.streetAddress was null in 94 of 107 measured records"
        ),
        "second_pass_must_be_independent": (
            "Same closed list, no sight of the first answer. Not a review of pass one."
        ),
        "reconciliation": {
            "both_same_zone": "accept, geo_verification_status = ai_agreement",
            "different_zones": "publication_zone_id = empty, reject_reason = ai_disagreement",
            "one_empty": "publication_zone_id = empty, reject_reason = zone_unknown",
            "both_empty": "publication_zone_id = empty, reject_reason = zone_unknown",
        },
        "review_trigger": (
            "If the weekly sample shows single-pass zone accuracy below 95% on "
            "explicit_address items, promote the mode to 'always'."
        ),
    },
    "lookup": {
        "keys": "name_display and aliases, normalised with the shared norm(): casefold, "
                "strip accents, punctuation to space, collapse runs of whitespace",
        "normalisation_note": (
            "One function, used by the build and by every consumer. v1.3 documented "
            "'strip punctuation, collapse spaces' and implemented neither collapsing nor "
            "the same rule on both sides: \"S. Angelo\" normalised to 's  angelo' with a "
            "double space, and 30 keys in the current gazetteer had the same problem."
        ),
        "index_includes_zone_names": (
            "The runtime index contains neighbourhood names, neighbourhood aliases AND "
            "publication-zone display names. The build checks ambiguity across all three, "
            "because a collision between a zone name and an unrelated neighbourhood would "
            "otherwise pass the build and surface as a silent coin flip at runtime (I7)."
        ),
        "ambiguous_keys": ["Colle del Sole"],
        "on_ambiguous": (
            "Never pick the first match. publication_zone_id = empty, "
            "reject_reason = ambiguous_alias."
        ),
        "zone_not_published_is_checked_here": (
            "An alias hit on Corcolle or Fiorano resolves to Agro Romano, which is not a "
            "page. That check belongs to the lookup step, not to post-model validation: the "
            "model is only ever given the 89 pages, so it can never return a non-page zone. "
            "Checking it only after the model leaves the real case unhandled."
        ),
        "containers_note": (
            "resolvable = false applies to the NEIGHBOURHOOD level only. Acilia, Ostia, "
            "Monteverde, Laurentino and Trigoria are containers as neighbourhoods and valid "
            "publication zones at the same time. A source that names one of them resolves to "
            "the zone with neighbourhood_id = null. Rejecting it would reject the commonest "
            "phrasing in Municipio X."
        ),
        "unpublished_zones_note": (
            "28 neighbourhoods map to 'Agro Romano' and one to 'Other'. Neither is a page. A "
            "correct alias hit on those produces a real place with nowhere to publish: "
            "reject_reason = zone_not_published, which is not the same failure as "
            "zone_not_in_list and must be counted separately."
        ),
        "free_text_note": (
            "Matching a gazetteer key inside a free-text description is not the same "
            "operation as matching it in location.name. See free_text_stoplist."
        ),
        "never_substring_note": (
            "A second, stricter list. free_text_stoplist protects descriptions only; "
            "never_substring protects every field, because an address carries its own city "
            "and country. Measured on the real sample: 'Italia' is a real gazetteer row "
            "publishing to Nomentano, and matching it inside addresses put 30 of 107 events "
            "on that one page. An exact whole-field match still resolves normally — "
            "location.name == 'Italia' is a legitimate hit; ', Italia' at the end of a "
            "street address is not."
        ),
    },
}

# Toponyms that are also ordinary Italian words, price tokens or common surnames. Matching
# these inside a description produces confident false positives — "nuovi talenti", "EUR 15",
# "prati fioriti" — and those false positives land in the one number that decides whether 89
# pages are executable. Safe in location.name / streetAddress / address, where the string is
# a place by construction. Not safe in a title or a description.
FREE_TEXT_STOPLIST = [
    "prati", "talenti", "eur", "omo", "ripa", "marconi", "morena",
    "centro storico", "trionfale", "quartaccio",
]

# Keys that must not be matched as a SUBSTRING anywhere — not in a description and not in
# a location field either. An exact whole-field match still resolves them normally.
#
# Found by running zone_distribution.py on the real 107-event sample, which is why the
# sample now ships with the package. "Italia" is a legitimate gazetteer row in Municipio II
# publishing to Nomentano — and every Italian postal address ends in ", Italia" or
# ", Italy". It matched 30 of 107 records and made Nomentano the busiest zone in Rome by a
# factor of six. v1.4 assumed a location field "is a place by construction" and applied no
# stoplist there; an address contains its own country, so that assumption was wrong.
NEVER_SUBSTRING = [
    "italia", "italy", "roma", "rome", "rm", "lazio",
]

GUARDS = {
    "note": (
        "Tier A runs before anything else and asks whether the coordinate is usable. "
        "Tier B runs AFTER the venue registry and asks whether the record is suspicious — "
        "a question a registry hit already answers. See resolution.order_of_operations."
    ),
    "tier_a": ["rome_bbox", "centroid_blocklist", "place_type_city"],
    "tier_b": ["name_coordinate_mismatch", "country_code_null_with_coordinates"],
    "rome_bbox": {
        "tier": "A",
        "min_lat": 41.6, "max_lat": 42.1, "min_lng": 12.15, "max_lng": 12.9,
        "action": "coordinates present and outside the box -> reject_reason = outside_rome, "
                  "regardless of what the model returns",
    },
    "centroid_blocklist": {
        "tier": "A",
        "note": (
            "Coordinates observed on unrelated events. They represent the city, not a venue. "
            "Resolving them yields a confident wrong answer in the historic centre. Extend "
            "this list whenever a coordinate pair recurs across unrelated events."
        ),
        "coordinates": [
            [41.900859832764, 12.483275413513],
            [41.9009311, 12.5012052],
            [41.9, 12.5],
            [41.882726005483, 12.490425109863],
        ],
        "tolerance_degrees": 1e-6,
        "action": "publication_zone_id = empty, reject_reason = centroid_detected",
    },
    "place_type_city": {
        "tier": "A",
        "action": "location.placeType == 'CITY' -> publication_zone_id = empty, "
                  "reject_reason = place_type_city. The city is known, the area is not.",
    },
    "name_coordinate_mismatch": {
        "tier": "B",
        "note": (
            "Observed live: records whose location.name read 'London, U.K.', 'Rim Park, "
            "Room 207' or 'Florence, Tuscany, Italy' while carrying Rome coordinates. "
            "Facebook backfills the search city's coordinates when it cannot resolve the "
            "venue."
        ),
        "v1_3_defect": (
            "v1.3 read the text after the last comma of location.name as a geographic tail "
            "and rejected anything not on an allowlist. But location.name is a VENUE name, "
            "not a hierarchical address: the tail is usually a room, a floor or a street "
            "number. On review the rule rejected 'Teatro Argentina, Sala Squarzina' and "
            "'Casa del Municipio, Sala Consiliare' — both perfectly valid Rome records. "
            "A guard that fires on valid records is worse than no guard, because it fails "
            "silently and irreversibly."
        ),
        "rule_location_city": {
            "applies_to": "location.city only",
            "why": (
                "location.city IS composite and hierarchical — measured: 'Rome, NY, United "
                "States', 'Brighton and Hove, United Kingdom'. The tail after the last "
                "comma is genuinely a region or a country."
            ),
            "test": "normalised tail must be in accepted_tails, or be a gazetteer key",
            "accepted_tails": [
                "roma", "rome", "rm", "italia", "italy", "lazio",
                "roma italia", "rome italy", "roma rm", "roma lazio",
                "citta metropolitana di roma capitale",
            ],
        },
        "rule_location_name": {
            "applies_to": "location.name",
            "test": (
                "Split the normalised name on commas. Reject only if a WHOLE segment "
                "equals a foreign_place_markers entry. Not a substring test: 'London Pub' "
                "is a plausible Rome venue and must pass, while 'London, U.K.' and a bare "
                "'Waterloo' must not. The negative test in test_guards.py caught this — a "
                "substring rule was the first fix attempted and it was the same defect one "
                "level down."
            ),
            "why_a_blocklist_here": (
                "The inverse of the city rule, and deliberately so. Venue names are "
                "unbounded, so an allowlist over them rejects the valid majority. A "
                "blocklist fails OPEN — an unlisted foreign city publishes — which is the "
                "acceptable direction for a name field, because the centroid blocklist, the "
                "bounding box and the countryCode guard also cover this case. Extend the "
                "list when a foreign toponym reaches a page."
            ),
            "foreign_place_markers": [
                "u k", "uk", "united kingdom", "england", "scotland", "wales", "ireland",
                "united states", "usa", "u s a", "canada", "ontario", "waterloo",
                "london", "paris", "berlin", "madrid", "barcelona", "amsterdam",
                "brussels", "vienna", "prague", "budapest", "athens", "lisbon",
                "new york", "boston", "denver", "brighton", "culver city", "bellaire",
                "wayland", "france", "germany", "spain", "portugal", "netherlands",
                "belgium", "austria", "switzerland", "greece",
                "florence", "firenze", "tuscany", "toscana", "milano", "milan",
                "napoli", "naples", "torino", "turin", "venezia", "venice",
                "bologna", "palermo", "genova", "genoa",
            ],
            "italian_markers_note": (
                "Italian cities other than Rome are on the list because "
                "'Florence, Tuscany, Italy' was observed carrying Rome coordinates and its "
                "country tail is legitimately 'italy'. The city rule cannot catch it."
            ),
            "note": (
                "'Rome' is deliberately NOT a marker: the keyword-search sample returned "
                "Rome NY and Rome GA, but those records are caught by the bounding box, and "
                "listing 'rome' would reject most legitimate Rome venues."
            ),
        },
        "action": "publication_zone_id = empty, reject_reason = name_coord_mismatch",
        "review_note": (
            "Both halves of this guard must be checked against VALID records, not only "
            "against the three anomalies that motivated it. The v1.3 package verified that "
            "every guard fired and never that one did not fire wrongly."
        ),
    },
    "country_code_null_with_coordinates": {
        "tier": "B",
        "note": (
            "Two of the three observed name/coordinate contradictions also carried "
            "countryCode = null. It is a usable suspicion signal — and in v1.2 it had an "
            "action but no reject code, so it was uncountable in the weekly sort."
        ),
        "action": "location.countryCode is null and coordinates are present -> "
                  "publication_zone_id = empty, reject_reason = country_code_missing",
        "tier_b_note": (
            "Runs after the venue registry from v1.4. A venue already resolved by a person "
            "is not made suspicious by a missing country code."
        ),
    },
}

# The local-interest filter. Structural signals narrow the field, the model judges what is
# left, nothing is deleted. Rationale and prompt text in EDITORIAL_FILTER.md.
EDITORIAL = {
    "note": (
        "Recurrence is a symptom, not the axis. A pub's Thursday karaoke and the same pub's "
        "one-off gig by a named band are both recurring-venue content, and only one of them "
        "is an event. The discriminator is whether the item has a reason to exist beyond "
        "the ordinary trading of whoever hosts it."
    ),
    "field": "editorial_class",
    "class_values": {
        "local_interest": "Publishable. An identified thing happening on a date.",
        "commercial_routine": "A real happening, but the ordinary programming of a "
                              "commercial venue, with no identified content.",
        "promotion": "Not an event. A product or a service being sold.",
        "uncertain": "The classifier could not decide. Queue item, not a discard.",
    },
    "criteria_in_execution_order": [
        {
            "step": 0,
            "test": "explicit_promotion",
            "weight": "gate — runs FIRST and stops",
            "question": "Is a product or a service being sold, with the event as framing? "
                        "Photo shoots, guided tours, cruises, discount offers, 'book your "
                        "session', 'limited places, contact us'.",
            "note": "Classify promotion and stop.",
            "v1_3_defect": (
                "v1.3 ran this test LAST, after identified_content. Most promotions carry "
                "exactly what test 1 looks for — a company, a title, a theme. 'Rome Sunset "
                "Photography Experience by XYZ' stopped at test 1 as local_interest and "
                "never reached the exclusion. The ordering neutralised the filter's whole "
                "purpose."
            ),
        },
        {
            "step": 1,
            "test": "identified_content",
            "weight": "dominant, among items that pass the gate",
            "question": "Is there a proper noun or a specific subject — a band, an author, "
                        "a speaker, a title, a theme?",
            "note": "If yes, classify local_interest. A named band playing on Friday in a "
                    "neighbourhood pub qualifies — free entry or not, and whether or not "
                    "that pub books a band every Friday.",
        },
        {
            "step": 1.5,
            "test": "community_organiser",
            "weight": "sufficient on its own",
            "question": "Is the organiser a community body — a neighbourhood committee, a "
                        "parish, an association, a school, a library, a social centre, the "
                        "municipio?",
            "note": "If yes, classify local_interest even when the content is generic. "
                    "'Assemblea del comitato di quartiere' and 'mercatino della "
                    "parrocchia' name no band and no author, and under the v1.4 tests "
                    "alone a monthly one could have been read as routine. This is the "
                    "basis value organizer_type, which v1.4 listed in the enum without "
                    "ever saying when to use it.",
        },
        {
            "step": 2,
            "test": "substitutability",
            "weight": "secondary",
            "question": "Absent identified content: is this item distinguishable from the "
                        "same venue's item last week?",
            "note": "'Karaoke', 'aperitivo con DJ', 'live music' with no act named — no. "
                    "Classify commercial_routine.",
        },
    ],
    "why_the_gate_is_first": (
        "A named seller is still a seller. Identified content distinguishes an occasion "
        "from routine programming; it cannot distinguish an occasion from a product, "
        "because products have names too. The two tests answer different questions and the "
        "exclusion has to run before the inclusion."
    ),
    "price_is_not_a_criterion": (
        "Free does not make an item an event and a ticket does not disqualify one. A "
        "five-euro concert is a concert. Do not reintroduce paidContent as a signal — it "
        "was false on all 187 measured records, including a ticketed concert."
    ),
    "recurrence_is_a_signal_not_a_gate": (
        "A weekly farmers' market and a monthly committee meeting are among the most "
        "valuable items on the site, so recurrence cannot be a gate. Whether a high "
        "occurrence_count leans commercial is an UNTESTED HYPOTHESIS, not a finding: the "
        "sample has 11 recurring series in 107 records, 8 WEEKLY and 3 CUSTOM, and nobody "
        "has classified them. v1.3 stated a threshold of ~20 and a community range of 8-12 "
        "dates; neither number came from the data and both are withdrawn. Record "
        "occurrence_count, correlate it against the human decisions in the review archive "
        "after a few weeks, and only then decide whether it is a usable prior."
    ),
    "basis_field": "editorial_basis",
    "basis_values": [
        "identified_content", "organizer_type", "venue_pattern",
        "explicit_promotion", "no_signal",
    ],
    "basis_definitions": {
        "identified_content": "Passed on a proper noun or a specific subject (step 1).",
        "organizer_type": "Passed because the organiser is a community body, whatever the "
                          "content (step 1.5).",
        "venue_pattern": "Classified commercial_routine on substitutability (step 2).",
        "explicit_promotion": "Excluded by the gate (step 0).",
        "no_signal": "Nothing to go on. Normally accompanies uncertain.",
    },
    "basis_note": (
        "Why the classifier decided, one value. It makes the queue sortable by cause and, "
        "after a few weeks, shows which judgements are reliable enough to promote to a "
        "deterministic rule — the same mechanism by which the gazetteer fattens on errors."
    ),
    "separate_pass_from_geography": (
        "Two judgements with different error rates must be countable separately (I5). Run "
        "the editorial classification as its own call, not merged into the geography "
        "prompt, or a drop in the weekly number cannot be attributed to either."
    ),
    "cross_block_contract": (
        "Corrected in v1.4. v1.3's Vision and Handover claimed editorial_class and "
        "content_type were already the common contract for all four page blocks and were "
        "not retrofittable. They are not: both enums are events-specific — a municipal "
        "resolution is not 'commercial_routine' and a news item is not a 'course'. What is "
        "genuinely common and genuinely not retrofittable is narrower and worth stating "
        "exactly: every item carries a geographic level (geo_level + municipality_code), a "
        "freshness field, and SOME editorial classification whose value set belongs to its "
        "block. The field names are shared; the enums extend per block, and the events "
        "enums below are simply the first of them."
    ),
    "content_type_field": "content_type",
    "content_type_values": {
        "event": "A single occasion.",
        "recurring_appointment": "A standing appointment of community value — a farmers' "
                                 "market, a committee meeting, a reading group.",
        "course": "Enrolment-based and multi-session: a language course, a theatre "
                  "workshop. Neither promotion nor an event in the ordinary sense. How it "
                  "is displayed is undecided; what matters today is that it is not "
                  "discarded indistinguishably in the meantime.",
    },
    "calibration": (
        "Before the filter goes live: hand-classify 30 items from the 107-event sample, run "
        "the prompt on the same 30, read the disagreements. They say whether the criterion "
        "is written clearly, and they are the examples to put in the prompt."
    ),
    "start_wide": (
        "Start permissive and tighten on evidence. In a low-density area a strict filter "
        "empties the block, and a wrongly filtered event is invisible — unlike a wrong "
        "zone, which is at least on a page where someone can see it."
    ),
    "asymmetry_warning": (
        "This filter fails silently in the direction that matters. The reject queue shows "
        "what it excluded, never what it excluded WRONGLY. That is what the ten sampled "
        "discards per week are for — see review.weekly_sample."
    ),
}

REVIEW = {
    "location": "one Google Sheet: one tab for the queue, one for the archive",
    "why_one_tab": (
        "Four blocks in four tabs means four schemas that diverge on the first field added "
        "to one of them, four n8n write nodes and four read-back workflows. One tab with a "
        "`block` column plus filter views gives the same working experience and one of each."
    ),
    "queue_columns": [
        "content_id", "block", "source_channel", "title", "excerpt", "venue_name",
        "location_id", "source_id", "matched_neighbourhood_id", "when", "reject_reason",
        "editorial_class", "editorial_basis", "ai_zone", "geo_basis", "source_url",
        "queued_at",
        "decision", "decided_zone", "fix_target", "fix_key", "fix_note",
        "decided_at", "decided_by",
    ],
    "editable_columns": ["decision", "decided_zone", "fix_target", "fix_key", "fix_note"],
    "decision_values": ["publish", "discard", "fix_zone", "skip"],
    "fix_target_values": ["gazetteer", "venue", "source", "none"],
    "fix_target_note": (
        "Four generic values, not fifteen block-specific ones. A dropdown you have to read "
        "is a dropdown that slows the review down."
    ),
    "structural_keys_note": (
        "Added in v1.4. v1.3 claimed the deciding reason for Sheets was that the fix "
        "applies in the same gesture with no bridge to build — and then carried no "
        "structured target for two of the four fix_target values. Manual review worked; "
        "deterministic write-back did not. Each target now has a key to write against:"
    ),
    "fix_target_keys": {
        "venue": "location_id (already present) + decided_zone -> upsert venues.json",
        "gazetteer": "matched_neighbourhood_id, or the zone from decided_zone when no "
                     "neighbourhood matched; fix_key holds the alias string to add",
        "source": "source_id — the Facebook page or group id — + decided_zone. This is the "
                  "page registry the posts channel will need, and it starts filling now.",
        "none": "no key. The decision is about this item only.",
    },
    "fix_key_note": (
        "One free-form column that means something different per fix_target: the alias to "
        "add, or nothing when the id columns already carry the key. fix_note stays prose "
        "for the human reader."
    ),
    "rules": [
        "n8n appends. Never sort or reorder the queue tab in place: the read-back matches "
        "on content_id, and manual reordering is how that breaks silently.",
        "Decided rows are moved to the archive tab by the workflow, never deleted by hand. "
        "The archive is the only record of human judgement in the system, and the only "
        "dataset against which the editorial filter can ever be evaluated.",
        "decided_at distinguishes 'not yet decided' from 'decided and not yet written "
        "back'. Without it a failed read-back is invisible.",
        "Every decision asks a second question: does this fix belong in the gazetteer, the "
        "venue registry or the source registry? A fix applied to the item alone solves one "
        "item.",
    ],
    "weekly_sample": {
        "published": 20,
        "discarded": 10,
        "published_question": "Is the zone right?",
        "discarded_question": "Was this genuinely not worth publishing?",
        "where_the_discarded_live": (
            "Corrected in v1.4. v1.3 said both 'nothing is deleted, an excluded item goes "
            "to the queue' and 'wrongly discarded items never appear at all', which cannot "
            "both be true. They can be reconciled and this is the reconciliation:"
        ),
        "discarded_sources": {
            "in_the_queue": "editorial_uncertain — a person sees these anyway",
            "auto_excluded_log": (
                "commercial_routine, promotion, and items skipped by a venue "
                "editorial_override = blacklist. These do NOT enter the working queue — "
                "putting hundreds of routine items in front of the reviewer would destroy "
                "the twenty-minute budget — and they are written to the archive tab with "
                "decision = auto_excluded. That log is the store the 10 are sampled from."
            ),
        },
        "note": (
            "The working queue shows false negatives the system chose to surface. It "
            "cannot show wrongly published items, which do not block because the system is "
            "confident, nor wrongly auto-excluded ones, which never reach it. The two "
            "samples are the only sensors for either."
        ),
    },
    "definition_of_done": {
        "block": "facebook_events",
        "it_is_a_stop_rule_not_a_proof": (
            "Corrected in v1.4. The minimum that satisfies the criteria is 19/20 twice, "
            "and 38/40 is compatible with a true accuracy anywhere between roughly 84% and "
            "99% at 95% confidence. Zero misses in 20 sampled discards likewise does not "
            "exclude a real error rate near 15%. Twenty a week is what one reviewer can "
            "actually do, so the sample size is not the thing to change — the claim is. "
            "This is a defensible rule for deciding to move on, not evidence that the "
            "resolver reached 95%. Do not report it as the latter."
        ),
        "criteria": [
            "two consecutive non-August weeks with zone accuracy at or above 95% on the "
            "20-item published sample",
            "the top reject reason stable across those two weeks, and understood",
            "no item in the discarded sample judged wrongly excluded two weeks running",
        ],
        "note": (
            "Written down before the work starts, because otherwise the decision to move on "
            "to the posts channel gets taken on a feeling. August is excluded on purpose: "
            "the 107-event sample was collected in the deadest week of the Italian year."
        ),
    },
}

# blocks_publication distinguishes a reject from a flag. v1.3 had three codes in this enum
# that do not stop publication (no_image, duplicate_suspected) or had no producing rule at
# all, which made the weekly sort mix two different kinds of thing.
REJECT_REASONS = [
    {"code": "outside_rome", "layer": "guard", "blocks_publication": True,
     "meaning": "Coordinates fall outside the Rome bounding box."},
    {"code": "centroid_detected", "layer": "guard", "blocks_publication": True,
     "meaning": "Coordinates match a known city-centroid pair, not a venue."},
    {"code": "place_type_city", "layer": "guard", "blocks_publication": True,
     "meaning": "location.placeType == CITY. City known, area unknowable."},
    {"code": "name_coord_mismatch", "layer": "guard", "blocks_publication": True,
     "meaning": "The record names a place that contradicts its Rome coordinates. Two "
                "rules: the comma tail of location.city is not an accepted Rome tail, or a "
                "whole comma-delimited segment of location.name matches a foreign place "
                "marker. See guards.name_coordinate_mismatch — one definition, not two."},
    {"code": "country_code_missing", "layer": "guard", "blocks_publication": True,
     "meaning": "location.countryCode is null with coordinates present. Suspicious, not "
                "publishable without review. New in v1.3 — in v1.2 the guard existed with "
                "no code, so it was uncountable."},
    {"code": "venue_unresolved", "layer": "lookup", "blocks_publication": True,
     "meaning": "TERMINAL, not a step-3 miss. Fires only when location.id is present, is "
                "not in the registry, AND the alias lookup and the model both failed to "
                "resolve the item. A registry miss on its own is not a reject: it simply "
                "continues to the next step. v1.3's WORKFLOW_FIXES said 'a miss is "
                "venue_unresolved', which — with location.id present in 106 of 107 records "
                "— would have queued almost every event before the alias lookup or the "
                "model ever ran."},
    {"code": "ambiguous_alias", "layer": "lookup", "blocks_publication": True,
     "meaning": "The matched name or alias belongs to more than one neighbourhood "
                "(currently: Colle del Sole)."},
    {"code": "zone_not_published", "layer": "lookup", "blocks_publication": True,
     "meaning": "Resolved to a real zone that is not a page (Agro Romano, Other). The place "
                "is right; there is nowhere to publish it. Fires in the lookup, not after "
                "the model — the model is only ever given the 89 pages."},
    {"code": "zone_not_in_list", "layer": "validation", "blocks_publication": True,
     "meaning": "The model returned a value that is not in the allowed list."},
    {"code": "ai_disagreement", "layer": "validation", "blocks_publication": True,
     "meaning": "The two independent passes returned different zones."},
    {"code": "zone_unknown", "layer": "validation", "blocks_publication": True,
     "meaning": "No zone could be determined, or an inferred zone had no stated location."},
    {"code": "date_unparseable", "layer": "extraction", "blocks_publication": True,
     "meaning": "duration could not be parsed and no end date could be derived."},
    {"code": "no_description", "layer": "extraction", "blocks_publication": True,
     "meaning": "description is null or under 40 characters after stripping whitespace. "
                "Measured: present on 100% of the direct-URL sample, so this should be rare "
                "and a rise in it means the extraction broke, not the source."},
    {"code": "no_image", "layer": "extraction", "blocks_publication": False,
     "meaning": "No image on the source record — roughly 10% of posts. FLAG ONLY: the item "
                "publishes without an image. v1.3 left this code with no producing rule at "
                "all, so it was a dead counter."},
    {"code": "duplicate_suspected", "layer": "extraction", "blocks_publication": False,
     "meaning": "Soft key location.id + utcStartDate matched an existing item. FLAG ONLY — "
                "flag, never delete, and never block. It appears in the queue for a human "
                "to merge or dismiss."},
    {"code": "not_an_event", "layer": "editorial", "blocks_publication": True,
     "meaning": "Explicit promotion: a product or service being sold with an event as "
                "framing. Photo shoots, tours, cruises. Rule in EDITORIAL_FILTER.md — in "
                "v1.2 this code existed with no rule that could produce it."},
    {"code": "commercial_routine", "layer": "editorial", "blocks_publication": True,
     "meaning": "A real happening with no identified content — the ordinary programming of "
                "a commercial venue. New in v1.3."},
    {"code": "editorial_uncertain", "layer": "editorial", "blocks_publication": True,
     "meaning": "The classifier returned uncertain. A queue item, not a discard. "
                "New in v1.3."},
]

VENUE_REGISTRY = {
    "location": "venues.json (exported) / Firestore collection `venues` (write path)",
    "key": "Facebook location.id",
    "not_here_because": (
        "Venue records are learned at runtime. gazetteer.json is regenerated from the "
        "workbook; merging the two would erase every learned venue on the next rebuild."
    ),
    "venue_nature_note": (
        "venue_nature is a signal passed to the editorial classifier, not a shortcut past "
        "it. A neighbourhood pub hosts both ordinary programming and real gigs, and a "
        "permanent verdict on the venue would discard the gigs. Only the extremes "
        "shortcut: an explicit whitelist (parish, committee, library, association) and a "
        "blacklist earned by repeated explicit_promotion."
    ),
}

# ------------------------------------------------------------------ BUILD ----


def sheet_dicts(wb, name):
    ws = wb[name]
    hdr = [c.value for c in ws[1]]
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        out.append({h: row[i] for i, h in enumerate(hdr) if h})
    return out


def as_bool(v):
    if isinstance(v, bool):
        return v
    return str(v).strip().upper() == "TRUE"


def as_str(v):
    return "" if v is None else str(v).strip()


def write_italian_export(wb, hoods, zones, strings, missing):
    """Regenerate the read-only Italian export.

    The MASTER sheet has claimed since v1.2 that this happens. Until v1.3 it did not, and
    the two workbooks had already diverged once — on the name of the key column."""
    def tr(s, context=None, record=True):
        """English -> Italian via the translation memory.

        `context` allows one English token to have two Italian renderings: the status value
        `added` is `aggiunto` on a neighbourhood row and `aggiunta` on a correction to a
        `row`. `record=False` is for columns that hold proper nouns rather than prose —
        place names are identical in both languages and must not be reported as missing."""
        s = as_str(s)
        if not s:
            return ""
        if context and f"{context}:{s}" in strings:
            return strings[f"{context}:{s}"]
        if s in strings:
            return strings[s]
        if record:
            missing.add(s)
        return s

    out = Workbook()
    ws = out.active
    ws.title = "AVVISO"
    for i, line in enumerate([
        f"EXPORT — NON MODIFICARE — v{VERSION}", "",
        "Questo file è un export in italiano. Il master è Rome_Neighbourhood_Gazetteer_EN.xlsx.",
        "Le modifiche fatte qui vengono perse alla prima rigenerazione.", "",
        "Per modificare i dati: edita il workbook EN e lancia  python build_gazetteer.py",
        "Per correggere una traduzione: edita it_strings.json e rilancia lo stesso comando.", "",
        "La colonna chiave si chiama 'neighbourhood_id', come nel master e nel JSON.",
    ], start=1):
        ws.cell(row=i, column=1, value=line)
    ws["A1"].font = Font(name="Arial", bold=True, size=14)
    ws.column_dimensions["A"].width = 95

    widths = (24, 26, 30, 8, 10, 24, 22, 10, 12, 12, 12, 40)

    def table(title, header, rows_):
        s = out.create_sheet(title)
        s.append(header)
        for c in s[1]:
            c.font = Font(name="Arial", bold=True)
        for r in rows_:
            s.append(r)
        s.freeze_panes = "A2"
        for col, w in zip("ABCDEFGHIJKL", widths):
            s.column_dimensions[col].width = w
        return s

    table("gazetteer",
          ["neighbourhood_id", "name_display", "aliases", "municipality_code",
           "spans_municipalities", "publication_zone", "publication_zone_id", "zu_code",
           "population", "resolvable", "status", "note"],
          [[h["neighbourhood_id"], h["name_display"], h["aliases"] or None,
            h["municipality_code"] or None, h["spans_municipalities"] or None,
            h["publication_zone"], h["publication_zone_id"], h["zu_code"] or None,
            h["population"], "TRUE" if h["resolvable"] else "FALSE",
            tr(h["status"]) or None, tr(h["note"]) or None] for h in hoods])

    table("publication_zones",
          ["publication_zone_id", "name_display", "primary_municipality",
           "spans_municipalities", "is_page", "search_queries", "note"],
          [[z["publication_zone_id"], z["name_display"], z["primary_municipality"] or None,
            z["spans_municipalities"] or None, "TRUE" if z["is_page"] else "FALSE",
            z["search_queries"] or None, tr(z["note"]) or None] for z in zones])

    table("correzioni", ["id", "campo", "prima", "dopo", "motivo"],
          [[c["id"], tr(c["field"]),
            tr(c["before"], context=c["field"], record=False) or c["before"],
            tr(c["after"], context=c["field"], record=False) or c["after"],
            tr(c["reason"])]
           for c in sheet_dicts(wb, "corrections")])

    table("da_verificare", ["id", "questione", "proposta"],
          [[q["id"], tr(q["issue"]), tr(q["proposal"])]
           for q in sheet_dicts(wb, "open_questions")])

    leg = out.create_sheet("legenda")
    for row in wb["legend"].iter_rows(values_only=True):
        leg.append([tr(v) if isinstance(v, str) else v for v in row])
    leg.column_dimensions["A"].width = 30
    leg.column_dimensions["B"].width = 100
    for row in leg.iter_rows():
        for c in row:
            c.font = Font(name="Arial", bold=(c.row == 2 or c.row == 8))
        row[-1].alignment = Alignment(wrap_text=True, vertical="top")

    out.save(EXPORT_IT)


def main():
    if not MASTER.exists():
        sys.exit(f"master workbook not found: {MASTER}")
    wb = load_workbook(MASTER, data_only=True)
    strings = json.loads(STRINGS.read_text(encoding="utf-8")) if STRINGS.exists() else {}

    zones = []
    for z in sheet_dicts(wb, "publication_zones"):
        zones.append({
            "publication_zone_id": as_str(z["publication_zone_id"]),
            "name_display": as_str(z["name_display"]),
            "primary_municipality": as_str(z["primary_municipality"]),
            "spans_municipalities": as_str(z["spans_municipalities"]),
            "is_page": as_bool(z["is_page"]),
            "search_queries": as_str(z["search_queries"]),
            "note": as_str(z["note"]),
        })
    by_name = {z["name_display"]: z for z in zones}

    hoods = []
    for n in sheet_dicts(wb, "gazetteer"):
        zone_name = as_str(n["publication_zone"])
        zone = by_name.get(zone_name)
        if zone is None:
            sys.exit(f"unjoinable publication_zone {zone_name!r} on {n['neighbourhood_id']!r}")
        zid = as_str(n.get("publication_zone_id")) or zone["publication_zone_id"]
        if zid != zone["publication_zone_id"]:
            sys.exit(f"publication_zone_id disagrees with the join on {n['neighbourhood_id']!r}")
        hoods.append({
            "neighbourhood_id": as_str(n["neighbourhood_id"]),
            "name_display": as_str(n["name_display"]),
            "aliases": as_str(n["aliases"]),
            "municipality_code": as_str(n["municipality_code"]),
            "spans_municipalities": as_str(n["spans_municipalities"]),
            "publication_zone_id": zid,
            "publication_zone": zone_name,
            "publishable": zone["is_page"],
            "zu_code": as_str(n["zu_code"]),
            "population": n["population"],
            "resolvable": as_bool(n["resolvable"]),
            "status": as_str(n["status"]),
            "note": as_str(n["note"]),
        })

    # integrity checks — a build that violates an invariant must fail, not warn
    ids = [h["neighbourhood_id"] for h in hoods]
    assert len(ids) == len(set(ids)), "duplicate neighbourhood_id"
    zids = [z["publication_zone_id"] for z in zones]
    assert len(zids) == len(set(zids)), "duplicate publication_zone_id"

    pages = [z["name_display"] for z in zones if z["is_page"]]

    # Ambiguity across the FULL runtime index: neighbourhood names, aliases AND zone display
    # names. v1.2 checked the first two only, so a zone name colliding with an unrelated
    # neighbourhood would have passed the build and become a coin flip at runtime (I7).
    # normalised with norm(), the SAME function the runtime index uses. v1.3 compared with
    # casefold() here while zone_distribution.py stripped accents and punctuation, so the
    # build's guarantee that a runtime collision fails the build was not actually true:
    # "Città Test" and "Citta Test" were two keys for the build and one at runtime.
    seen = {}
    for h in hoods:
        keys = [h["name_display"]] + [a.strip() for a in h["aliases"].split("|") if a.strip()]
        for k in keys:
            seen.setdefault(norm(k), set()).add(h["publication_zone_id"])
    for z in zones:
        seen.setdefault(norm(z["name_display"]), set()).add(z["publication_zone_id"])
    found = {k for k, v in seen.items() if len(v) > 1}
    declared = {norm(a) for a in RESOLUTION["lookup"]["ambiguous_keys"]}
    if found != declared:
        sys.exit(
            "ambiguous lookup keys changed.\n"
            f"  in data:     {sorted(found)}\n"
            f"  declared:    {sorted(declared)}\n"
            "Resolve the duplicate, or update RESOLUTION['lookup']['ambiguous_keys']."
        )

    # every stoplist entry must actually be a lookup key, or it is silently protecting
    # nothing and hiding a typo
    unknown = [s for s in FREE_TEXT_STOPLIST if norm(s) not in seen]
    if unknown:
        sys.exit(f"FREE_TEXT_STOPLIST entries that are not lookup keys: {unknown}")
    # NEVER_SUBSTRING may legitimately contain non-keys (roma, italy): they are listed so
    # that adding such a row later does not silently reintroduce the defect.

    doc = {
        "version": VERSION,
        "source": "Rome_Neighbourhood_Gazetteer_EN.xlsx",
        "generated_by": "build_gazetteer.py — do not hand-edit",
        "neighbourhoods": hoods,
        "publication_zones": zones,
        "resolution": RESOLUTION,
        "free_text_stoplist": FREE_TEXT_STOPLIST,
        "never_substring": NEVER_SUBSTRING,
        "guards": GUARDS,
        "editorial": EDITORIAL,
        "review": REVIEW,
        "reject_reasons": REJECT_REASONS,
        "venue_registry": VENUE_REGISTRY,
    }
    JSON_OUT.write_text(json.dumps(doc, ensure_ascii=False, indent=1) + "\n", encoding="utf-8")

    header = (
        "PUBLICATION ZONES — allowed values for the `publication_zone` field.\n"
        f"{len(pages)} values. Use the exact spelling below, and nothing else on the line.\n"
        "If the source does not let you determine the zone, return an empty string.\n"
        "An empty value is a correct answer; a plausible guess is not.\n"
        "\n"
        "GENERATED by build_gazetteer.py from publication_zones where is_page = true.\n"
        "Do not edit by hand.\n\n"
    )
    PROMPT_OUT.write_text(header + "\n".join(sorted(pages)) + "\n", encoding="utf-8")

    missing = set()
    write_italian_export(wb, hoods, zones, strings, missing)

    print(f"gazetteer.json      {len(hoods)} neighbourhoods, {len(zones)} zones "
          f"({len(pages)} pages)")
    print(f"prompt list         {len(pages)} values")
    print(f"italian export      {EXPORT_IT}")
    print(f"not published       "
          f"{sum(1 for h in hoods if not h['publishable'])} neighbourhoods map to a non-page zone")
    print(f"free-text stoplist  {len(FREE_TEXT_STOPLIST)} keys not matched inside descriptions")
    print(f"never-substring     {len(NEVER_SUBSTRING)} keys matched only on a whole field")
    if missing:
        print(f"\nUNTRANSLATED        {len(missing)} string(s) copied through in English.")
        print("                    Add them to it_strings.json and re-run:")
        for s in sorted(missing)[:10]:
            print(f"                      {s[:88]}")


if __name__ == "__main__":
    main()
